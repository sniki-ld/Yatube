from openpyxl import load_workbook
from ..models import Group


def xlsx_group_parser(file):
    wb = load_workbook(file, read_only=True, data_only=True)
    sheet_main = wb.active
    groups = []
    for row in sheet_main.iter_rows(min_row=2):
        title, slug, description = [cell.value for cell in row]
        group = Group(title=title, slug=slug, description=description)
        groups.append(group)
        return groups
